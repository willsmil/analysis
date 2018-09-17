from openpyxl import Workbook, load_workbook


#打开exel
def open_exel(resource):
    wb = load_workbook(resource)
    return wb


#获取每行数据
def get_row_data(sheet, row_num):
    data_list = []
    for cell in list(sheet.rows)[row_num]:
        value = cell.value
        if value == None:
            continue
        data_list.append(value)
    return data_list


def get_data_start():
    wb1 = open_exel('..\\resource\\201806.xlsx')
    print(wb1.get_sheet_names())
    ws = wb1.get_sheet_by_name('Sheet2')
    sheet = wb1.active
    listAll = []
    for i in range(24, 80):
        list = get_row_data(sheet, i)
        print("i, list1 length: ", i, len(list))
        if len(list) == 56:
            listAll.append(list)
    return listAll


def get_data_end():
    wb1 = open_exel('..\\resource\\201809.xlsx')
    print(wb1.get_sheet_names())
    ws = wb1.get_sheet_by_name('Sheet2')
    sheet = wb1.active
    listAll = []
    for i in range(24, 85):
        list = get_row_data(sheet, i)
        print("i, list2 length: ", i, len(list))
        if len(list) == 55:
            listAll.append(list)
    return listAll


#找到两表中都有的职能
def process():
    write = Workbook()
    sheet = write.active
    sheet.title = "Sheet1"
    sheet.append(["职能", "total change", "increase", "decrease"])
    list1 = get_data_end()
    list2 = get_data_start()
    print("list1: ", list1)
    print("list2: ", list2)
    ret_list = list(set(list1[0]).union(set(list2[0])))
    for i in ret_list:
        print(1, i)
        listpos1 = []
        if i in list1[0]:
            index1 = list1[0].index(i)
            for v in list1:
                if v[index1] == "X":
                    listpos1.append(v[0]+v[1])
        print(2, listpos1)

        listpos2 = []
        if i in list2[0]:
            index2 = list2[0].index(i)
            for v in list2:
                if v[index2] == "X":
                    listpos2.append(v[0] + v[1])
        print(3, listpos2)
        increase = list(set(listpos1).difference(set(listpos2)))
        increasestr = '\n'.join(increase)
        print("increase: ", increase)
        decrease = list(set(listpos2).difference(set(listpos1)))
        decreasestr = '\n'.join(decrease)
        print("decrease: ", decrease)
        totalchange = len(increase) - len(decrease)
        print("total change", totalchange)
        sheet.append([i, totalchange, increasestr, decreasestr])
    write.save('..\\resource\\result09.xlsx')


#找到两表中不同的职能
def find_other():
    list1 = get_data_end()
    list2 = get_data_start()
    other = list(set(list1[0]).difference(set(list2[0])))
    print(other)
    other = list(set(list2[0]).difference(set(list1[0])))
    print(other)


if __name__ == '__main__':
    process()